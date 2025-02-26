import re, json
from datetime import date, timedelta
from openpyxl import load_workbook

cls = {
    "第一节": {"start": "08:00", "end": "08:45"},
    "第二节": {"start": "08:50", "end": "09:35"},
    "第三节": {"start": "09:55", "end": "10:40"},
    "第四节": {"start": "10:45", "end": "11:30"},
    "第五节": {"start": "11:35", "end": "12:20"},
    "第六节": {"start": "14:00", "end": "14:45"},
    "第七节": {"start": "14:50", "end": "15:35"},
    "第八节": {"start": "15:40", "end": "16:25"},
    "第九节": {"start": "16:45", "end": "17:30"},
    "第十节": {"start": "17:35", "end": "18:20"},
    "第十一节": {"start": "19:00", "end": "19:45"},
    "第十二节": {"start": "19:50", "end": "20:35"},
    "第十三节": {"start": "20:40", "end": "21:25"}
}

def read_excel_to_2d_list(file_path, start_cell='C8'):
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active

    start_row = sheet[start_cell].row
    start_col = sheet[start_cell].column

    data_list = []
    for row in sheet.iter_rows(min_row=start_row, min_col=start_col, max_row=39, max_col=9, values_only=True):
        for index, cell in enumerate(row):
            if cell:
                for item in cell.split('（本）'):
                    if item.split():
                        data_list.append([i for i in ','.join(item.split()).split(',') if i])
    return data_list

def format_tap_json(day, Data, notice=30):
    res = []
    for _, item in Data.items():
        for ds in item['days']:
            it = {
                "title": item['title'],
                "start": cls[item['classes'][0]]['start'],
                "end": cls[item['classes'][1]]['end'],
                "day": str(day + timedelta(days=ds)),
                "gmt": "+8",
                "notice": notice,
                "color": "0xFFFFBFBF",
                "locate": item['locate'],
                "data": ""
            }
            res.append(it)
    return sorted(res, key=lambda x: (x['day'], x['start']))

file_path = 'timeTableForStu12.xlsx'
classData = read_excel_to_2d_list(file_path, start_cell='C8')
classDict = {}

for i, data in enumerate(classData):
    title = data[0].split('-')[1].split('[')[0]
    days = []
    for s in data[1:-3]:
        st = int(data[-3][-1])
        if '-' in s:
            match = re.search(r"(\d+)-(\d+)周", s)
            d = list(range(int(match.group(1)) - 1, int(match.group(2))))
            days.extend([x * 7 + st for x in d])
        else:
            days.append(st + int(s[:-1]) * 7)
    classes = data[-2].split('-')
    locate = data[-1]

    classDict[i] = {
        'title': title,
        'days': days,
        'classes': classes,
        'locate': locate
    }

classJson = format_tap_json(date(2025, 2, 23), classDict, notice=30)

with open('classJson.json', 'w', encoding='utf-8') as file:
    json.dump({"notice": classJson}, file, indent=4, ensure_ascii=False)
